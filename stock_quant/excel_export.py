"""把漲幅排行 (list[RankRow]) 寫進一個 .xlsx 檔。

用 openpyxl;每次呼叫覆蓋整個檔案 (對應「每次覆蓋同一檔」)。
寫檔失敗 (例如檔案正被 Excel 開著鎖住) 不會中斷主迴圈，改回傳錯誤字串。
"""
from __future__ import annotations

from datetime import datetime
from typing import Optional, Sequence

_HEADERS = ["排名", "代號", "名稱", "市場", "收盤/現價", "漲跌", "漲幅%", "成交量(張)",
            "開盤", "最高", "最低", "昨收"]


def save_ranking(path: str, now: datetime, rows: Sequence, source: str = "live",
                 title: str = "漲幅排行") -> Optional[str]:
    """把排行寫進 path。成功回傳 None，失敗回傳錯誤訊息字串。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "未安裝 openpyxl (pip install openpyxl) -> 略過寫 Excel"

    tag = "盤中即時" if source == "live" else "最後交易日"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "漲幅排行"

        # 第一列: 快照標題
        ws.append([f"{title}  {now:%Y-%m-%d %H:%M:%S}  [{tag}]  共 {len(rows)} 檔"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))

        # 第二列: 表頭
        ws.append(_HEADERS)
        head_fill = PatternFill("solid", fgColor="DDEBF7")
        for col in range(1, len(_HEADERS) + 1):
            c = ws.cell(row=2, column=col)
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")

        # 資料列
        for i, r in enumerate(rows, start=1):
            ws.append([
                i, r.symbol, r.name, r.market.zh,
                r.close, r.change, r.change_pct, r.lots,
                r.open, r.high, r.low, r.prev_close,
            ])

        # 數字格式
        first_data = 3
        last_data = first_data + len(rows) - 1
        if len(rows) > 0:
            for row in ws.iter_rows(min_row=first_data, max_row=last_data):
                row[6].number_format = "0.00"      # 漲幅%
                for cell in (row[4], row[5], row[8], row[9], row[10], row[11]):
                    cell.number_format = "0.00"    # 價格欄
                row[7].number_format = "#,##0"     # 成交量(張)

        # 欄寬 + 凍結表頭
        widths = [6, 8, 12, 6, 11, 9, 9, 13, 10, 10, 10, 10]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "A3"

        wb.save(path)
        return None
    except PermissionError:
        return f"無法寫入 {path} (檔案可能正被 Excel 開著) -> 本次略過"
    except Exception as exc:  # noqa: BLE001 — 寫檔失敗不可中斷排行主迴圈
        return f"寫 Excel 失敗: {type(exc).__name__}: {exc}"
