"""把回測結果寫成 Excel 績效報告 (openpyxl)。

工作表：摘要 / 交易明細 / 每日權益(含權益曲線圖) / 每日選股(起漲篩選明細)。
"""
from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if TYPE_CHECKING:
    from .engine import BacktestResult

_HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
_WIN_FILL = PatternFill("solid", fgColor="E2EFDA")
_LOSS_FILL = PatternFill("solid", fgColor="FCE4E4")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def _header(ws, row, headers):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.fill, c.font, c.alignment, c.border = _HEAD_FILL, _HEAD_FONT, _CENTER, _BORDER


def _autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _summary_sheet(ws, result):
    cfg = result.config
    ws["A1"] = "台股起漲策略 回測績效摘要"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = (f"回測區間：{cfg.start} ~ {cfg.end}　|　選股：漲幅{cfg.min_change_pct:g}%池 + "
                f"起漲6條件(紅K/突破昨高/量增/站上5MA/站上月線上彎/昨在5MA下) + 強度分排序")
    ws["A3"] = (f"庫存上限 {cfg.max_holdings} 檔，部位={cfg.sizing_mode}，"
                f"出場={cfg.ma_exit}MA跌破" + (f" 或停損{cfg.stop_loss_pct:g}%" if cfg.stop_loss_pct > 0 else "")
                + "；漲停不買/跌停不賣；進出場用當日收盤價")
    ws["A3"].font = Font(italic=True, size=9, color="808080")
    r = 5
    _header(ws, r, ["績效指標", "數值"])
    r += 1
    for k, v in result.metrics.items():
        ws.cell(row=r, column=1, value=k).border = _BORDER
        cell = ws.cell(row=r, column=2, value=v)
        cell.border = _BORDER
        if isinstance(v, (int, float)):
            if "率" in k or "%" in k:
                cell.number_format = "0.00"
            elif any(t in k for t in ("資金", "權益", "損益", "獲利", "虧損", "成本")):
                cell.number_format = "#,##0"
        if k == "總損益" and isinstance(v, (int, float)):
            cell.fill = _WIN_FILL if v >= 0 else _LOSS_FILL
        r += 1
    _autosize(ws, {"A": 18, "B": 20})


def _trades_sheet(ws, result):
    headers = ["代號", "市場", "進場日", "進場價", "出場日", "出場價", "股數", "持有天數",
               "出場原因", "毛損益", "成本", "淨損益", "報酬率%", "強度分", "進場理由"]
    _header(ws, 1, headers)
    for i, t in enumerate(result.trades, start=2):
        row = [t.symbol, t.market, t.entry_date, t.entry_price, t.exit_date, t.exit_price,
               t.shares, t.holding_days, t.exit_reason, t.gross_pnl, t.cost, t.pnl,
               t.return_pct, t.score, t.note]
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = _BORDER
            if j in (4, 6, 13, 14):
                c.number_format = "0.00"
            elif j in (10, 11, 12):
                c.number_format = "#,##0"
        fill = _WIN_FILL if t.pnl > 0 else (_LOSS_FILL if t.pnl < 0 else None)
        if fill:
            for j in range(1, len(headers) + 1):
                ws.cell(row=i, column=j).fill = fill
    ws.freeze_panes = "A2"
    _autosize(ws, {"A": 8, "B": 7, "C": 12, "D": 9, "E": 12, "F": 9, "G": 8, "H": 9,
                   "I": 12, "J": 12, "K": 10, "L": 12, "M": 10, "N": 8, "O": 40})


def _equity_sheet(ws, result):
    _header(ws, 1, ["日期", "權益", "現金", "持股檔數"])
    for i, (d, eq, cash, n) in enumerate(result.equity_curve, start=2):
        ws.cell(row=i, column=1, value=d).border = _BORDER
        for col, val, fmt in ((2, round(eq, 0), "#,##0"), (3, round(cash, 0), "#,##0"), (4, n, "0")):
            c = ws.cell(row=i, column=col, value=val)
            c.number_format = fmt
            c.border = _BORDER
    n_rows = len(result.equity_curve)
    if n_rows >= 2:
        chart = LineChart()
        chart.title = "權益曲線 (Equity Curve)"
        chart.y_axis.title = "權益 (TWD)"
        chart.x_axis.title = "日期"
        chart.height, chart.width = 10, 26
        data = Reference(ws, min_col=2, min_row=1, max_row=n_rows + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F2")
    ws.freeze_panes = "A2"
    _autosize(ws, {"A": 12, "B": 14, "C": 14, "D": 10})


def _selection_sheet(ws, result):
    _header(ws, 1, ["日期", "排名", "代號", "名稱", "現價", "漲幅%", "量比",
                    "5MA", "月線20MA", "月線上彎", "強度分", "理由"])
    r = 2
    for d, picks in result.daily_selection:
        for rank, sr in enumerate(picks, start=1):
            row = sr.row
            vals = [d, rank, row.symbol, row.name, row.close, row.change_pct, sr.vol_ratio,
                    sr.ma5, sr.ma20, "↑" if sr.ma20_up else "", sr.score, " / ".join(sr.reasons)]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = _BORDER
                if j in (5, 6, 7, 8, 9):
                    c.number_format = "0.00"
                elif j == 11:
                    c.number_format = "0.0"
            r += 1
    ws.freeze_panes = "A2"
    _autosize(ws, {"A": 12, "B": 5, "C": 8, "D": 12, "E": 9, "F": 8, "G": 7,
                   "H": 9, "I": 10, "J": 8, "K": 8, "L": 40})


def write_report(result, path: str) -> str:
    wb = Workbook()
    _summary_sheet(wb.active, result)
    wb.active.title = "摘要"
    _trades_sheet(wb.create_sheet("交易明細"), result)
    _equity_sheet(wb.create_sheet("每日權益"), result)
    _selection_sheet(wb.create_sheet("每日選股"), result)
    wb.save(path)
    return path
