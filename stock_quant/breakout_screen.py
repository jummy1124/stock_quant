"""起漲點篩選: 接在 IntradayRanker.tick() 之後的第二道篩選。

上游 IntradayRanker 已篩出「漲幅 3% ~ 漲停前一檔」的個股池 (list[RankRow])。
本模組對這個池子再用以下 6 個硬條件過濾 (全部通過才入選)：

  1. 當日K棒為紅K            : 現價/收盤 > 開盤
  2. 突破前一交易日最高點    : 現價/收盤 > 昨高
  3. 量增 (預設 1.2 倍)       : 當日成交量 > 前一交易日成交量 × 1.2
  4. 站上五日均價            : 現價/收盤 > 5MA
  5. 站上月均線且月均線上彎  : 現價/收盤 > 20MA 且 20MA 向上 (今日 20MA > N 日前 20MA)
  6. 昨日仍在五日線下        : 前一交易日收盤 < 前一交易日的 5MA (今日才轉強上穿，鎖定當日起漲)

通過 6 條件的個股即入選，依評估順序輸出 (不另做排序)。

資料直接來自 ranker (RankRow + ranker.history())，不經 Excel 中轉。
歷史K不足以算某條均線時，該條件自動略過 (不誤殺)；預設 --days 已確保足夠。

⚠️ 篩選為資訊參考，非投資建議。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, time
from typing import Optional, Sequence


# ============================================================
# 參數設定 (照盤性微調都改這裡)
# ============================================================

@dataclass
class BreakoutConfig:
    # --- 條件門檻 ---
    vol_ratio_min: float = 1.2             # 條件3: 當日量 / 昨量 下限 (預設 1.2 倍)
    ma_short: int = 5                      # 條件4: 五日均價 (5MA)
    ma_mid: int = 20                       # 條件5: 月均線 (20MA)
    ma_mid_slope_lookback: int = 5         # 條件5: 月均線「向上」的回看天數 (今日20MA > N日前20MA)

    # 量能比較方式: False=當日累積量直接比昨量 (照規格，早盤較嚴)；
    #               True =先把當日量換算成全日預估量再比 (盤中早盤較公允)
    use_volume_projection: bool = False

    # --- 台股交易時段 (use_volume_projection=True 時換算用) ---
    session_start: time = time(9, 0)
    session_end: time = time(13, 30)


@dataclass
class ScoredRow:
    """一檔通過 6 條件的個股 + 明細。"""
    row: object                  # 原始 RankRow
    prev_high: Optional[float]   # 昨日最高 (條件2 基準)
    vol_ratio: Optional[float]   # 當日量 / 昨量
    ma5: Optional[float]
    ma20: Optional[float]
    ma20_up: bool                # 月均線是否上彎
    reasons: list = field(default_factory=list)   # 入選理由 (印出用)


# ============================================================
# 工具
# ============================================================

def _elapsed_fraction(now: datetime, cfg: BreakoutConfig) -> float:
    """已經過的交易時間比例 (0.01~1)，把當日累積量換算成全日預估量用。"""
    t = now.time()
    s, e = cfg.session_start, cfg.session_end
    if t <= s:
        return 0.01
    if t >= e:
        return 1.0
    total = (e.hour * 60 + e.minute) - (s.hour * 60 + s.minute)
    passed = (t.hour * 60 + t.minute) - (s.hour * 60 + s.minute)
    return max(passed / total, 0.01)


def _ma(closes: list, n: int) -> Optional[float]:
    """最近 n 根收盤均線；K 棒不足回傳 None。"""
    if n <= 0 or len(closes) < n:
        return None
    return sum(closes[-n:]) / n


def _ma_closes(ranker, row) -> list:
    """供『站上均線』(5MA/月線) 比較用的收盤序列 —— 只取『已收盤的完成日K』。

    live (盤中即時): 歷史到昨日為止。**不把盤中尚未收盤的今日現價算進均線**，
        月線/5MA 以完成日K為準 (等同券商顯示的「月均價」)。先前版本會把今日現價
        append 進去再取最後 20 根，等於把最舊一根擠掉、又用今日價墊高/拉低自己的均線，
        對「已從高點回落」的個股會算出偏低的月線，導致現價其實在月線下卻誤判「站上月線」。
    eod  (最後交易日): 歷史最後一根本來就是今日的完成收盤，已包含在內。

    比較時再用今日現價 (row.close) 去和這條均線比，才是真正的『站上均線』。
    """
    hist = ranker.history(row.symbol) or []
    return [float(q.close) for q in hist if q.close is not None]


def _prev_high_and_vol(ranker, row):
    """取『前一交易日』的最高價與成交量 (條件2 基準 + 條件3 分母)。"""
    hist = ranker.history(row.symbol) or []
    if getattr(ranker, "last_source", "live") == "live":
        prev = hist[-1] if hist else None          # live: 最後一根=昨日完成日K
    else:
        prev = hist[-2] if len(hist) >= 2 else None  # eod: 倒數第二根=昨日
    ph = float(prev.high) if prev and prev.high is not None else None
    pv = float(prev.volume) if prev and prev.volume is not None else None
    return ph, pv


def _closes_through_yesterday(ranker, row) -> list:
    """收盤序列『到前一交易日為止』(不含今日)，給條件6算昨日的 5MA 用。

    live (盤中即時): 歷史本身就到昨日 -> 直接用
    eod  (最後交易日): 歷史最後一根是今日 -> 去掉最後一根
    """
    hist = ranker.history(row.symbol) or []
    h_closes = [float(q.close) for q in hist if q.close is not None]
    if getattr(ranker, "last_source", "live") == "live":
        return h_closes
    return h_closes[:-1]


# ============================================================
# 核心: 6 條件評估
# ============================================================

def _evaluate(ranker, row, now, cfg: BreakoutConfig) -> Optional[ScoredRow]:
    """逐一檢查 6 條件，全過回傳 ScoredRow，否則 None。歷史不足以算的均線條件自動略過。"""
    price = float(row.close) if row.close is not None else None
    op = float(row.open) if getattr(row, "open", None) is not None else None
    if price is None:
        return None

    closes = _ma_closes(ranker, row)            # 均線用完成日K (盤中不含今日現價)
    prev_high, prev_vol = _prev_high_and_vol(ranker, row)
    reasons = []

    # 條件1: 紅K (現價/收盤 > 開盤)
    if op is None:
        reasons.append("紅K(無開盤價,略過)")
    elif price > op:
        reasons.append("紅K")
    else:
        return None

    # 條件2: 突破前一交易日最高點
    if prev_high is None:
        return None                          # 沒有昨高就無法判斷起漲核心，剔除
    if price > prev_high:
        reasons.append(f"突破昨高{prev_high:g}")
    else:
        return None

    # 條件3: 當日量 > 昨量 × 1.2
    vol_ratio = None
    if prev_vol and getattr(row, "volume", None) is not None:
        today_vol = float(row.volume)
        if cfg.use_volume_projection:
            today_vol /= _elapsed_fraction(now, cfg)
        vol_ratio = today_vol / prev_vol
        if vol_ratio >= cfg.vol_ratio_min:
            reasons.append(f"量比{vol_ratio:.2f}")
        else:
            return None
    else:
        reasons.append("量能(無昨量/今量,略過)")

    # 條件4: 站上五日均價 (5MA)
    ma5 = _ma(closes, cfg.ma_short)
    if ma5 is None:
        reasons.append("5MA(歷史不足,略過)")
    elif price > ma5:
        reasons.append("站上5MA")
    else:
        return None

    # 條件5: 站上月均線且月均線上彎 (20MA)
    ma20 = _ma(closes, cfg.ma_mid)
    lb = cfg.ma_mid_slope_lookback
    ma20_prev = _ma(closes[:-lb], cfg.ma_mid) if lb > 0 and len(closes) > cfg.ma_mid + lb else None
    ma20_up = False
    if ma20 is None:
        reasons.append("月線(歷史不足,略過)")
    else:
        if price <= ma20:
            return None
        if ma20_prev is None:
            reasons.append("站上月線(斜率不足,略過上彎判斷)")
        elif ma20 > ma20_prev:
            ma20_up = True
            reasons.append("站上月線↑")
        else:
            return None                      # 月線未上彎

    # 條件6: 前一交易日收盤 < 前一交易日的 5MA (昨日仍在五日線下，今日才轉強上穿)
    prev_seq = _closes_through_yesterday(ranker, row)
    ma5_prev = _ma(prev_seq, cfg.ma_short)
    prev_close = float(row.prev_close) if getattr(row, "prev_close", None) is not None \
        else (prev_seq[-1] if prev_seq else None)
    if ma5_prev is None or prev_close is None:
        reasons.append("昨日5MA(歷史不足,略過)")
    elif prev_close < ma5_prev:
        reasons.append("昨收<昨日5MA")
    else:
        return None                          # 昨日已在五日線上 -> 非當日起漲

    return ScoredRow(row=row, prev_high=prev_high,
                     vol_ratio=(round(vol_ratio, 2) if vol_ratio is not None else None),
                     ma5=(round(ma5, 2) if ma5 is not None else None),
                     ma20=(round(ma20, 2) if ma20 is not None else None),
                     ma20_up=ma20_up, reasons=reasons)


# ============================================================
# 對外主函式
# ============================================================

def screen_breakout(ranker, rows: Sequence, now: Optional[datetime] = None,
                    cfg: Optional[BreakoutConfig] = None) -> list:
    """對 ranker.tick() 的結果做 6 條件起漲篩選，回傳 ScoredRow 清單 (依評估順序)。"""
    cfg = cfg or BreakoutConfig()
    now = now or datetime.now()
    out: list[ScoredRow] = []
    for row in rows:
        sr = _evaluate(ranker, row, now, cfg)
        if sr is not None:
            out.append(sr)
    return out


# ============================================================
# 輸出 Excel (可選)
# ============================================================

_HEADERS = ["排名", "代號", "名稱", "市場", "現價", "漲幅%", "昨高", "量比",
            "5MA", "月線(20MA)", "月線上彎", "理由"]


def save_breakout(path: str, now: datetime, scored: Sequence,
                  source: str = "live") -> Optional[str]:
    """把起漲篩選結果寫進 path。成功回傳 None，失敗回傳錯誤字串 (不中斷主迴圈)。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "未安裝 openpyxl -> 略過寫起漲 Excel"

    tag = "盤中即時" if source == "live" else "最後交易日"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "起漲篩選"
        ws.append([f"起漲個股 (紅K+突破昨高+量增+站上5MA+月線上彎+昨日在5MA下)  {now:%Y-%m-%d %H:%M:%S}  [{tag}]  共 {len(scored)} 檔"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))

        ws.append(_HEADERS)
        fill = PatternFill("solid", fgColor="FCE4D6")
        for col in range(1, len(_HEADERS) + 1):
            c = ws.cell(row=2, column=col)
            c.font = Font(bold=True)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")

        for i, sr in enumerate(scored, start=1):
            r = sr.row
            ws.append([i, r.symbol, r.name, r.market.zh, r.close, r.change_pct,
                       sr.prev_high, sr.vol_ratio, sr.ma5, sr.ma20,
                       "↑" if sr.ma20_up else "", " / ".join(sr.reasons)])

        # 數字格式 (資料從第 3 列開始: 1=標題, 2=表頭)
        first = 3
        last = first + len(scored) - 1
        if scored:
            for row in ws.iter_rows(min_row=first, max_row=last):
                for c in (row[4], row[6], row[8], row[9]):   # 現價 / 昨高 / 5MA / 月線
                    c.number_format = "0.00"
                row[5].number_format = "0.00"                # 漲幅%
                row[7].number_format = "0.00"                # 量比
                row[10].alignment = Alignment(horizontal="center")  # 月線上彎

        # 欄寬 + 凍結標題與表頭
        widths = [6, 8, 12, 6, 9, 8, 9, 7, 9, 11, 9, 40]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "A3"

        wb.save(path)
        return None
    except PermissionError:
        return f"無法寫入 {path} (檔案可能正被 Excel 開著) -> 本次略過"
    except Exception as exc:  # noqa: BLE001 — 寫檔失敗不可中斷主迴圈
        return f"寫起漲 Excel 失敗: {type(exc).__name__}: {exc}"
